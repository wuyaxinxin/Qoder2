# 始终生效
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import random
import string
import os

print('🚀 开始创建大型Excel文件...')

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = '大数据表'

# 添加表头
headers = ['ID', '姓名', '年龄', '性别', '邮箱', '电话', '地址', '城市', '部门', '职位', 
           '入职日期', '工资', '绩效评分', '项目经验', '技能标签', '备注信息', '状态', '更新时间']
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 生成大量数据(50000行以确保超过5M)
print('📝 正在生成数据...')
cities = ['北京', '上海', '广州', '深圳', '杭州', '南京', '成都', '武汉', '西安', '重庆']
departments = ['研发部', '市场部', '销售部', '人力资源部', '财务部', '运营部', '客服部', '产品部']
positions = ['工程师', '经理', '主管', '专员', '总监', '助理', '顾问', '分析师']
statuses = ['在职', '离职', '试用期', '实习']

for i in range(1, 50001):
    name = ''.join(random.choices(string.ascii_uppercase, k=2)) + '员工' + str(i)
    age = random.randint(22, 60)
    gender = random.choice(['男', '女'])
    email = f'user{i}@company.com.cn.example.longdomainname'
    phone = f'1{random.randint(3,9)}{random.randint(100000000, 999999999)}'
    address = f'{random.choice(cities)}市某区某街道{random.randint(1,999)}号{random.randint(1,50)}栋{random.randint(1,20)}单元{random.randint(100,999)}室'
    city = random.choice(cities)
    dept = random.choice(departments)
    pos = random.choice(positions)
    hire_date = f'20{random.randint(15,24):02d}-{random.randint(1,12):02d}-{random.randint(1,28):02d}'
    salary = random.randint(5000, 50000)
    score = round(random.uniform(60, 100), 2)
    exp = f'参与项目{random.randint(1,20)}个,累计工作经验{random.randint(1,15)}年,主要负责系统开发、测试、部署等工作内容'
    skills = ','.join(random.choices(['Python', 'Java', 'JavaScript', 'C++', 'Go', 'React', 'Vue', 'Docker', 'Kubernetes', 'MySQL', 'Redis', 'MongoDB'], k=5))
    notes = '这是一段很长的备注信息,用于增加文件大小。' * 3 + f'员工编号{i}的详细信息记录。'
    status = random.choice(statuses)
    update_time = f'2024-{random.randint(1,12):02d}-{random.randint(1,28):02d} {random.randint(0,23):02d}:{random.randint(0,59):02d}:{random.randint(0,59):02d}'
    
    ws.append([i, name, age, gender, email, phone, address, city, dept, pos, 
               hire_date, salary, score, exp, skills, notes, status, update_time])
    
    if i % 5000 == 0:
        print(f'   已生成 {i:,} 行数据...')

# 调整列宽
print('📐 调整列宽...')
column_widths = [8, 12, 8, 8, 35, 15, 50, 10, 12, 12, 12, 10, 12, 60, 50, 80, 10, 20]
for idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

# 保存文件
filename = 'large_data.xlsx'
print(f'💾 正在保存文件 {filename}...')
wb.save(filename)
print(f'✅ Excel文件创建成功!')

# 检查文件大小
size_bytes = os.path.getsize(filename)
size_mb = size_bytes / (1024 * 1024)
print(f'\n📊 文件信息:')
print(f'   文件大小: {size_mb:.2f} MB ({size_bytes:,} 字节)')
print(f'   数据行数: 50,000 行 + 1 表头行')
print(f'   数据列数: 18 列')
print(f'   文件位置: {os.path.abspath(filename)}')

if size_mb >= 5:
    print(f'\n🎉 成功! 文件大小 {size_mb:.2f} MB > 5 MB')
else:
    print(f'\n⚠️  警告: 文件大小 {size_mb:.2f} MB < 5 MB')
